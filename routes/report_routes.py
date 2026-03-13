# from flask import Blueprint, request, jsonify, session
# from datetime import datetime
# from db import get_db

# report_bp = Blueprint("report", __name__)

# @report_bp.route("/reports", methods=["GET"])
# def generate_report():
#     user_id = session.get("user_id")
#     if not user_id:
#         return jsonify({"error": "Unauthorized"}), 401

#     period = request.args.get("period", "monthly") # monthly / yearly
#     month = request.args.get("month")              # 10
#     year = request.args.get("year")                # 2023

#     now = datetime.now()
#     month = month or str(now.month)
#     year = year or str(now.year)
#     try:
#         month_int = int(month)
#         year_int = int(year)
#     except (TypeError, ValueError):
#         return jsonify({"error": "Invalid month/year"}), 400

#     conn = get_db()
#     if conn is None:
#         return jsonify({"error": "Database connection failed"}), 500
#     cursor = conn.cursor(dictionary=True)

#     # ------------ TOTAL INCOME ------------
#     cursor.execute("""
#         SELECT SUM(amount) AS income 
#         FROM transactions 
#         WHERE user_id = %s AND type = 'Income'
#         AND MONTH(date) = %s AND YEAR(date) = %s
#     """, (user_id, month_int, year_int))
#     total_income = cursor.fetchone()["income"] or 0

#     # ------------ TOTAL EXPENSE ------------
#     cursor.execute("""
#         SELECT SUM(amount) AS expense 
#         FROM transactions 
#         WHERE user_id = %s AND type = 'Expense'
#         AND MONTH(date) = %s AND YEAR(date) = %s
#     """, (user_id, month_int, year_int))
#     total_expense = cursor.fetchone()["expense"] or 0

#     # ------------ CATEGORY BREAKDOWN ------------
#     cursor.execute("""
#         SELECT category, SUM(amount) AS total
#         FROM transactions
#         WHERE user_id = %s AND type='Expense'
#         AND MONTH(date) = %s AND YEAR(date) = %s
#         GROUP BY category
#     """, (user_id, month_int, year_int))
#     category_rows = cursor.fetchall()

#     category_breakdown = []
#     for row in category_rows:
#         percent = (row["total"] / total_expense * 100) if total_expense else 0
#         category_breakdown.append({
#             "category": row["category"],
#             "amount": row["total"],
#             "percent": round(percent, 2)
#         })

#     # ------------ DAILY TREND (BAR CHART) ------------
#     cursor.execute("""
#         SELECT DAY(date) AS day, SUM(amount) AS total
#         FROM transactions
#         WHERE user_id = %s AND type='Expense'
#         AND MONTH(date) = %s AND YEAR(date) = %s
#         GROUP BY day
#         ORDER BY day ASC
#     """, (user_id, month_int, year_int))
    
#     daily_rows = cursor.fetchall()
#     daily_trend = [{"day": r["day"], "amount": r["total"]} for r in daily_rows]

#     cursor.close()
#     conn.close()

#     return jsonify({
#         "period": period,
#         "month": month,
#         "year": year,
#         "total_income": total_income,
#         "total_expense": total_expense,
#         "net_balance": total_income - total_expense,
#         "categories": category_breakdown,
#         "daily_trend": daily_trend
#     })

from flask import Blueprint, request, jsonify, session, Response
from datetime import datetime
from io import StringIO, BytesIO
import csv
import calendar
from db import get_db

report_bp = Blueprint("report", __name__)

def _parse_month_year(args):
    now = datetime.now()
    month_raw = args.get("month")
    year_raw = args.get("year")

    try:
        month = int(month_raw) if month_raw else now.month
        year = int(year_raw) if year_raw else now.year
    except (TypeError, ValueError):
        return None, None, "Invalid month/year"

    if month < 1 or month > 12:
        return None, None, "Month must be between 1 and 12"
    if year < 1970 or year > 2100:
        return None, None, "Year out of range"

    return month, year, None


def _build_report(user_id, month, year):
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    # TOTAL INCOME
    cursor.execute("""
        SELECT SUM(amount) AS total
        FROM transactions
        WHERE user_id=%s AND LOWER(type)='income'
        AND MONTH(date)=%s AND YEAR(date)=%s
    """, (user_id, month, year))
    row = cursor.fetchone()
    total_income = row["total"] if row and row["total"] else 0

    # TOTAL EXPENSE
    cursor.execute("""
        SELECT SUM(amount) AS total
        FROM transactions
        WHERE user_id=%s AND LOWER(type)='expense'
        AND MONTH(date)=%s AND YEAR(date)=%s
    """, (user_id, month, year))
    row = cursor.fetchone()
    total_expense = row["total"] if row and row["total"] else 0

    # CATEGORY BREAKDOWN
    cursor.execute("""
        SELECT category, SUM(amount) AS amount
        FROM transactions
        WHERE user_id=%s AND LOWER(type)='expense'
        AND MONTH(date)=%s AND YEAR(date)=%s
        GROUP BY category
    """, (user_id, month, year))
    categories = cursor.fetchall()

    # DAILY TREND
    cursor.execute("""
        SELECT DAY(date) AS day, SUM(amount) AS amount
        FROM transactions
        WHERE user_id=%s AND LOWER(type)='expense'
        AND MONTH(date)=%s AND YEAR(date)=%s
        GROUP BY day
        ORDER BY day
    """, (user_id, month, year))
    daily_rows = cursor.fetchall()

    days_in_month = calendar.monthrange(year, month)[1]
    daily_map = {}
    for row in daily_rows:
        day = int(row.get("day") or 0)
        if day:
            daily_map[day] = float(row.get("amount") or 0)

    cursor.close()
    conn.close()

    daily = [{"day": day, "amount": daily_map.get(day, 0)} for day in range(1, days_in_month + 1)]

    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "net_balance": total_income - total_expense,
        "categories": categories,
        "daily_trend": daily,
    }


@report_bp.route("/api/reports", methods=["GET"])
def generate_report():

    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    month, year, err = _parse_month_year(request.args)
    if err:
        return jsonify({"error": err}), 400

    report = _build_report(user_id, month, year)
    return jsonify(report)


@report_bp.route("/api/reports/export", methods=["GET"])
def export_report():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    month, year, err = _parse_month_year(request.args)
    if err:
        return jsonify({"error": err}), 400

    fmt = (request.args.get("format") or "csv").lower()
    report = _build_report(user_id, month, year)

    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["Section", "Key", "Value"])
        writer.writerow(["Totals", "Total Income", report["total_income"]])
        writer.writerow(["Totals", "Total Expense", report["total_expense"]])
        writer.writerow(["Totals", "Net Balance", report["net_balance"]])
        writer.writerow([])
        writer.writerow(["Categories", "Category", "Amount"])
        for row in report["categories"]:
            writer.writerow(["Categories", row["category"], row["amount"]])
        writer.writerow([])
        writer.writerow(["Daily Trend", "Day", "Amount"])
        for row in report["daily_trend"]:
            writer.writerow(["Daily Trend", row["day"], row["amount"]])

        filename = f"report-{year:04d}-{month:02d}.csv"
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if fmt in ("xlsx", "excel"):
        try:
            from openpyxl import Workbook
        except ImportError:
            return jsonify({"error": "Excel export requires openpyxl (pip install openpyxl)."}), 501

        wb = Workbook()
        ws_totals = wb.active
        ws_totals.title = "Totals"
        ws_totals.append(["Total Income", report["total_income"]])
        ws_totals.append(["Total Expense", report["total_expense"]])
        ws_totals.append(["Net Balance", report["net_balance"]])

        ws_categories = wb.create_sheet("Categories")
        ws_categories.append(["Category", "Amount"])
        for row in report["categories"]:
            ws_categories.append([row["category"], row["amount"]])

        ws_daily = wb.create_sheet("Daily Trend")
        ws_daily.append(["Day", "Amount"])
        for row in report["daily_trend"]:
            ws_daily.append([row["day"], row["amount"]])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"report-{year:04d}-{month:02d}.xlsx"
        return Response(
            bio.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if fmt == "pdf":
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            return jsonify({"error": "PDF export requires reportlab (pip install reportlab)."}), 501

        bio = BytesIO()
        c = canvas.Canvas(bio, pagesize=letter)
        width, height = letter
        y = height - 50
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, y, f"Financial Report {year:04d}-{month:02d}")
        y -= 30
        c.setFont("Helvetica", 12)
        c.drawString(50, y, f"Total Income: {report['total_income']}")
        y -= 18
        c.drawString(50, y, f"Total Expense: {report['total_expense']}")
        y -= 18
        c.drawString(50, y, f"Net Balance: {report['net_balance']}")
        y -= 28
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Categories")
        y -= 16
        c.setFont("Helvetica", 11)
        for row in report["categories"]:
            c.drawString(60, y, f"{row['category']}: {row['amount']}")
            y -= 14
            if y < 60:
                c.showPage()
                y = height - 50
        y -= 10
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Daily Trend")
        y -= 16
        c.setFont("Helvetica", 11)
        for row in report["daily_trend"]:
            c.drawString(60, y, f"Day {row['day']}: {row['amount']}")
            y -= 14
            if y < 60:
                c.showPage()
                y = height - 50
        c.save()
        bio.seek(0)

        filename = f"report-{year:04d}-{month:02d}.pdf"
        return Response(
            bio.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return jsonify({"error": "Unsupported format"}), 400
